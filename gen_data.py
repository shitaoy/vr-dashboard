import openpyxl
import re
import json
from datetime import datetime

excel_path = r'F:\BaiduSyncdisk\08 2026年度\06 2026-勘-032号 2026年管内危岩落石、泥石流安全评估及航拍VR建模\02 评估\作品集1.xlsx'
output_path = r'C:\Users\28366\WorkBuddy\2026-07-24-11-42-59\js\data.js'

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['Sheet1']

works = []
tunnel_mileages = {}  # key: line + ":" + tunnelName

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=1):
    line = str(row[0]).strip() if row[0] else ''
    if not line:
        continue

    # 隧道编号：处理 "/" 和 None
    raw_num = row[1]
    if raw_num is not None and str(raw_num).strip() not in ('', '/'):
        try:
            tunnel_num = int(str(raw_num).strip())
        except ValueError:
            tunnel_num = 0
    else:
        tunnel_num = 0

    tunnel_name_raw = str(row[2]).strip() if row[2] else ''
    mileage_raw = str(row[3]).strip() if row[3] else ''

    # 经纬度：处理 None
    lng = float(row[4]) if row[4] is not None else 0.0
    lat = float(row[5]) if row[5] is not None else 0.0
    elevation = round(float(row[6]), 1) if row[6] is not None else 0.0

    workshop = str(row[7]).strip() if row[7] else ''

    # 拍摄时间 -> 年份
    if row[8] is not None:
        if isinstance(row[8], datetime):
            year = str(row[8].year)
        else:
            year = str(row[8])[:4]
    else:
        year = ''

    url = str(row[9]).strip() if row[9] else ''

    # 位置类型检测
    if '进口' in tunnel_name_raw:
        position = '进口'
    elif '出口' in tunnel_name_raw:
        position = '出口'
    elif '涵洞' in tunnel_name_raw:
        position = '涵洞'
    elif '中心点' in tunnel_name_raw:
        position = '中心点'
    else:
        position = '其他'

    # 清理隧道名称
    clean_name = tunnel_name_raw
    for suffix in ['（进口）', '（出口）', '(进口)', '(出口)', '（涵洞）', '(涵洞)']:
        clean_name = clean_name.replace(suffix, '')
    clean_name = clean_name.replace('中心点', '').strip()
    if not clean_name:
        clean_name = tunnel_name_raw

    # 清理里程
    mileage = mileage_raw
    for suffix in ['（进口）', '（出口）', '(进口)', '(出口)', '（涵洞）', '(涵洞)']:
        mileage = mileage.replace(suffix, '')
    mileage = mileage.strip()
    if mileage == '/':
        mileage = ''

    # 隧道里程范围（用 line:name 作为复合键）
    m_match = re.search(r'K0*(\d+)\+(\d+)', mileage, re.IGNORECASE)
    if m_match:
        m_val = int(m_match.group(1)) * 1000 + int(m_match.group(2))
        key = line + ':' + clean_name
        if key not in tunnel_mileages:
            tunnel_mileages[key] = {'min': m_val, 'max': m_val, 'name': clean_name, 'line': line}
        else:
            tunnel_mileages[key]['min'] = min(tunnel_mileages[key]['min'], m_val)
            tunnel_mileages[key]['max'] = max(tunnel_mileages[key]['max'], m_val)

    works.append({
        'id': i,
        'line': line,
        'tunnelNum': tunnel_num,
        'tunnelName': clean_name,
        'position': position,
        'mileage': mileage,
        'lng': round(lng, 6),
        'lat': round(lat, 6),
        'elevation': elevation,
        'workshop': workshop,
        'year': year,
        'url': url
    })

# 隧道汇总
tunnels_summary = []
for key in sorted(tunnel_mileages.keys()):
    info = tunnel_mileages[key]
    length = info['max'] - info['min']
    tunnels_summary.append({
        'line': info['line'],
        'name': info['name'],
        'length': length
    })

# 线路定义（按实际数据排序）
lines_list = [
    {'name': '杭温铁路', 'code': 'hw', 'color': '#00d4ff'},
    {'name': '杭昌高铁', 'code': 'hc', 'color': '#f59e0b'},
    {'name': '衢九线', 'code': 'qj', 'color': '#a855f7'},
    {'name': '沪昆线', 'code': 'hk', 'color': '#10b981'},
]

with open(output_path, 'w', encoding='utf-8') as f:
    f.write('/**\n')
    f.write(' * ============================================================\n')
    f.write(' *  浙江省高铁VR全景作品集 - 数据配置文件\n')
    f.write(' * ============================================================\n')
    f.write(' *  数据来源：作品集1.xlsx（自动生成）\n')
    f.write(' *  包含线路：杭温铁路、杭昌高铁、衢九线、沪昆线\n')
    f.write(' *  字段说明：\n')
    f.write(' *    line       - 铁路线名\n')
    f.write(' *    tunnelNum  - 隧道编号（无编号为0）\n')
    f.write(' *    tunnelName - 隧道名称（已去除进出口后缀）\n')
    f.write(' *    position   - 进口 / 出口 / 涵洞 / 中心点 / 其他\n')
    f.write(' *    mileage    - 里程桩号\n')
    f.write(' *    lng/lat    - 全景点经纬度（无坐标为0）\n')
    f.write(' *    elevation  - 全景点高程（米）\n')
    f.write(' *    workshop   - 所在维修车间\n')
    f.write(' *    year       - 拍摄年份\n')
    f.write(' *    url        - 720云作品链接\n')
    f.write(' * ============================================================\n')
    f.write(' */\n\n')
    f.write('const VR_DATA = {\n\n')
    f.write('  // ===== 铁路线路定义（颜色用于地图标注区分） =====\n')
    f.write('  lines: ' + json.dumps(lines_list, ensure_ascii=False, indent=4) + ',\n\n')
    f.write('  // ===== VR全景类型 =====\n')
    f.write('  categories: ["进口", "出口", "涵洞", "中心点", "其他"],\n\n')
    f.write('  // ===== 隧道汇总（线路、名称、长度米） =====\n')
    f.write('  tunnels: ' + json.dumps(tunnels_summary, ensure_ascii=False, indent=4) + ',\n\n')
    f.write('  // ===== VR全景作品列表（共 ' + str(len(works)) + ' 条） =====\n')
    f.write('  works: [\n')

    for w in works:
        # 转义字符串中的特殊字符
        def esc(s):
            return str(s).replace('\\', '\\\\').replace('"', '\\"')

        f.write('    {')
        f.write('id: ' + str(w['id']) + ', ')
        f.write('line: "' + esc(w['line']) + '", ')
        f.write('tunnelNum: ' + str(w['tunnelNum']) + ', ')
        f.write('tunnelName: "' + esc(w['tunnelName']) + '", ')
        f.write('position: "' + esc(w['position']) + '", ')
        f.write('mileage: "' + esc(w['mileage']) + '", ')
        f.write('lng: ' + str(w['lng']) + ', ')
        f.write('lat: ' + str(w['lat']) + ', ')
        f.write('elevation: ' + str(w['elevation']) + ', ')
        f.write('workshop: "' + esc(w['workshop']) + '", ')
        f.write('year: "' + esc(w['year']) + '", ')
        f.write('url: "' + esc(w['url']) + '"')
        f.write('},\n')

    f.write('  ]\n')
    f.write('};\n')

# 统计输出
lines_stat = {}
years_stat = {}
for w in works:
    lines_stat[w['line']] = lines_stat.get(w['line'], 0) + 1
    years_stat[w['year']] = years_stat.get(w['year'], 0) + 1

print(f'Generated data.js with {len(works)} VR works, {len(tunnels_summary)} tunnels')
print(f'Lines: {lines_stat}')
print(f'Years: {years_stat}')
print(f'File written to: {output_path}')
